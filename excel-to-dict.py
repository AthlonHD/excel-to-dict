# -*- coding=utf-8 -*-

import xlrd
from xlrd import xldate_as_tuple
import datetime
from openpyxl import load_workbook

'''
xlrd中单元格的数据类型
数字一律按浮点型输出，日期输出成一串小数，布尔型输出0或1，所以我们必须在程序中做判断处理转换
成我们想要的数据类型
0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
'''


class workExcel:

    def __init__(self, excel_path, sheet_name):
        self.excel_book = xlrd.open_workbook(r'%s' % excel_path)  # 读取excel，路径使用原生字符串
        self.excel_table = self.excel_book.sheet_by_name(sheet_name)  # 根据sheet名称读取
        print('Sheet读取检查：', self.excel_book.sheet_loaded(sheet_name))
        self.keys = self.excel_table.row_values(0)  # 读取行标题
        print('标题栏：', self.keys)
        self.nrows = self.excel_table.nrows  # 读取行
        self.ncols = self.excel_table.ncols  # 读取列
        print('行数：', self.nrows)
        print('列数：', self.ncols)

    def readData(self):     # 将表格内的数据读取为列表
        # 定义一个空列表
        table_data = []
        for i in range(1, self.nrows):
            # 定义一个空字典
            sheet_data = {}
            for j in range(self.ncols):
                # 获取单元格数据类型
                f_type = self.excel_table.cell(i, j).ctype
                # 获取单元格数据
                f_cell = self.excel_table.cell_value(i, j)
                if f_type == 2 and f_cell % 1 == 0:  # 如果是整形
                    f_cell = int(f_cell)
                elif f_type == 3:
                    # 转成datetime对象
                    date = datetime.datetime(*xldate_as_tuple(f_cell, 0))
                    f_cell = date.strftime('%Y-%m-%d')  # /%m %H:%M:%S
                elif f_type == 4:
                    f_cell = True if f_cell == 1 else False
                sheet_data[self.keys[j]] = f_cell
                # 循环每一个有效的单元格，将字段与值对应存储到字典中
                # 字典的key就是excel表中每列第一行的字段
                # sheet_data[self.keys[j]] = self.table.row_values(i)[j]
            # 再将字典追加到列表中
            table_data.append(sheet_data)
        # 返回从excel中获取到的数据：以列表存字典的形式返回
        return table_data

    def writeData(self):    # 写入到文件
        wb = load_workbook("% s" % excel_path)  # 生成一个已存在的workbook对象
        wb_active = wb.active  # 激活sheet

        wb.save("% s" % excel_path)  # 保存
        print('写入excel操作完成！')


if __name__ == '__main__':
    excel_path = r""     # 设置excel文件路径
    sheet_name = ""   # 输入sheet名称
    get_data = workExcel(excel_path, sheet_name)
    datas = get_data.readData()
    print('EXCEL读取列表：', datas)

    # get_data.writeData()    # 写入文件
